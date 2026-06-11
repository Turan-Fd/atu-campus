import argparse
import json
import re
from pathlib import Path

from openpyxl import load_workbook


def clean(value):
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return re.sub(r"\s+", " ", str(value)).strip()


def normalize_header(value):
    return clean(value).casefold()


def find_column(headers, *names):
    normalized_names = {normalize_header(name) for name in names}
    for index, header in enumerate(headers):
        if normalize_header(header) in normalized_names:
            return index
    return -1


def cell(row, index):
    return clean(row[index]) if index >= 0 and index < len(row) else ""


def read_students(workbook_path):
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    students = []

    for sheet_name, level in (("Bakalavriat", "Bakalavriat"), ("Magistr", "Magistr")):
        sheet = workbook[sheet_name]
        rows = sheet.iter_rows(values_only=True)
        headers = list(next(rows))
        columns = {
            "status": find_column(headers, "Təhsil alır"),
            "course": find_column(headers, "Kurs"),
            "work": find_column(headers, "İŞ NÖMRƏSİ", "İş nömrəsi"),
            "surname": find_column(headers, "SOYAD"),
            "name": find_column(headers, "AD"),
            "father": find_column(headers, "ATA ADI", "ATA adı"),
            "study": find_column(headers, "TƏHSİLALMA FORMASI", "Təhsil növü"),
            "faculty": find_column(headers, "FAKÜLTƏ"),
            "group": find_column(headers, "QRUP", "Qrup"),
            "specialty": find_column(headers, "İXTİSAS", "İxtisas"),
            "specialization": find_column(headers, "İxtisaslaşma"),
            "gender": find_column(headers, "CİNSİ/kişi/qadın", "Cinsi"),
            "identity_card": find_column(
                headers,
                "ŞƏXSİYYƏT VƏSİQƏSİ nömrəsi",
                "Şəxsiyyət vəsiqəsinin seriyası və nömrəsi",
            ),
            "fin": find_column(headers, "FİN"),
        }

        for row in rows:
            work_number = re.sub(r"\D", "", cell(row, columns["work"])).lstrip("0")
            if not 4 <= len(work_number) <= 10:
                continue
            students.append(
                {
                    "level": level,
                    "status": cell(row, columns["status"]),
                    "course": cell(row, columns["course"]),
                    "workNumber": work_number,
                    "surname": cell(row, columns["surname"]),
                    "name": cell(row, columns["name"]),
                    "fatherName": cell(row, columns["father"]),
                    "studyForm": cell(row, columns["study"]),
                    "faculty": cell(row, columns["faculty"]),
                    "group": cell(row, columns["group"]),
                    "specialty": cell(row, columns["specialty"]),
                    "specialization": cell(row, columns["specialization"]),
                    "gender": cell(row, columns["gender"]),
                    "identityCard": cell(row, columns["identity_card"]),
                    "fin": cell(row, columns["fin"]),
                }
            )

    return students


def build_photo_manifest(photo_directory):
    manifest = {}
    valid_extensions = {".jpg", ".jpeg", ".png", ".webp"}
    for photo in sorted(photo_directory.rglob("*")):
        if not photo.is_file() or photo.suffix.casefold() not in valid_extensions:
            continue
        if not photo.stem.isdigit():
            continue
        work_number = photo.stem.lstrip("0") or "0"
        manifest.setdefault(work_number, photo.relative_to(photo_directory).as_posix())
    return manifest


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--workbook", required=True, type=Path)
    parser.add_argument("--photos", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()

    students = read_students(args.workbook)
    photos = build_photo_manifest(args.photos)
    args.output.mkdir(parents=True, exist_ok=True)

    (args.output / "statistika-students.json").write_text(
        json.dumps(students, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    (args.output / "student-photos.json").write_text(
        json.dumps(photos, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )

    counts = {}
    for student in students:
        number = student["workNumber"]
        counts[number] = counts.get(number, 0) + 1

    duplicate_numbers = sorted(number for number, count in counts.items() if count > 1)
    matched_photos = sum(1 for number in counts if number in photos)
    print(f"Student records: {len(students)}")
    print(f"Unique work numbers: {len(counts)}")
    print(f"Duplicate work numbers: {len(duplicate_numbers)}")
    print(f"Indexed photos: {len(photos)}")
    print(f"Students with photo: {matched_photos}")


if __name__ == "__main__":
    main()
